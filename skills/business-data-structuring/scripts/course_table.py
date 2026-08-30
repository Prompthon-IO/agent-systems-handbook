"""Bounded CSV/XLSX parsing and explicit, deterministic table normalization."""
from __future__ import annotations
import csv
import datetime as dt
import re
from collections import Counter
from decimal import Decimal, InvalidOperation
from pathlib import Path

from course_runtime import CourseError, digest, file_hash

NULLS = {"", "null", "none", "n/a", "na"}


def text_value(value):
    if value is None:
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()
    return str(value).strip()


def read_table(path: Path, sheet: str | None = None) -> tuple[list[str], list[list[str]], dict]:
    if not path.is_file() or path.is_symlink() or path.stat().st_size > 20_000_000:
        raise CourseError("INVALID_DATASET", "Select a CSV/XLSX file smaller than 20 MB, not a symlink.")
    selected = None
    rows = []
    if path.suffix.lower() == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as stream:
            reader = csv.reader(stream)
            for row in reader:
                rows.append([text_value(v) for v in row])
                if len(rows) > 10001:
                    raise CourseError("DATASET_TOO_LARGE", "Classroom input is bounded to 10,000 rows.")
    elif path.suffix.lower() == ".xlsx":
        try:
            import openpyxl
        except ImportError:
            raise CourseError("DEPENDENCY_MISSING", "Install this package's requirements.txt in the course venv for XLSX input.") from None
        book = openpyxl.load_workbook(path, read_only=True, data_only=False, keep_links=False)
        try:
            if sheet is None and len(book.sheetnames) != 1:
                raise CourseError("SHEET_REQUIRED", "A multi-sheet workbook requires an explicit --sheet; do not guess the business table.")
            selected = sheet or book.sheetnames[0]
            if selected not in book.sheetnames:
                raise CourseError("INVALID_SHEET", "Selected worksheet does not exist.")
            for row in book[selected].iter_rows():
                if any(cell.data_type == "f" for cell in row):
                    raise CourseError("FORMULA_INPUT", "Formula cells need an explicitly reviewed values-only export. This helper does not trust stale cached results or execute formulas.")
                rows.append([text_value(c.value) for c in row])
                if len(rows) > 10001:
                    raise CourseError("DATASET_TOO_LARGE", "Classroom input is bounded to 10,000 rows.")
        finally:
            book.close()
    else:
        raise CourseError("UNSUPPORTED_FILE", "Only UTF-8 comma-separated CSV and .xlsx are supported.")
    if not rows or not rows[0] or len(rows[0]) > 100:
        raise CourseError("INVALID_DATASET", "Supply one header row with 1–100 columns.")
    headers = rows[0]
    body = [row for row in rows[1:] if any(v.strip() for v in row)]
    if any(len(row) != len(headers) for row in body):
        raise CourseError("RAGGED_ROWS", "Row widths differ; repair the source copy instead of silently dropping cells.")
    return headers, body, {"source_ref": path.name, "source_sha256": file_hash(path), "format": path.suffix.lower().lstrip("."), "sheet": selected}


def headers_normalized(headers: list[str]) -> list[str]:
    used, result = set(), []
    for index, header in enumerate(headers):
        base = re.sub(r"[^\w]+", "_", header.casefold()).strip("_") or f"column_{index + 1}"
        name, suffix = base, 2
        while name in used:
            name, suffix = f"{base}_{suffix}", suffix + 1
        used.add(name)
        result.append(name)
    return result


def number(value: str) -> Decimal:
    # Explicit English numeric convention. Never guess decimal-comma locales.
    if not re.fullmatch(r"[-+]?(?:\d+|\d{1,3}(?:,\d{3})+)(?:\.\d+)?", value):
        raise ValueError("not an unambiguous decimal number")
    parsed = Decimal(value.replace(",", ""))
    if not parsed.is_finite():
        raise ValueError("nonfinite")
    return parsed


def infer(values: list[str]) -> str:
    values = [v for v in values if v.casefold() not in NULLS]
    if not values:
        return "text"
    if all(re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", v) for v in values):
        return "email"
    if all(re.fullmatch(r"\d{4}-\d{2}-\d{2}", v) for v in values):
        return "date"
    try:
        if not any(re.fullmatch(r"0\d+", v) for v in values):
            for value in values:
                number(value)
            return "number"
    except (ValueError, InvalidOperation):
        pass
    return "text"


def normalize_cell(value: str, spec: dict):
    if value.casefold() in NULLS:
        return None
    kind = spec["type"]
    if kind == "text":
        return value
    if kind == "email":
        if not re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", value):
            raise ValueError("invalid_email")
        return value.casefold()
    if kind == "date":
        return dt.datetime.strptime(value, spec.get("format", "%Y-%m-%d")).date().isoformat()
    if kind == "number":
        return format(number(value), "f")
    if kind == "currency":
        currency = spec.get("currency")
        if not isinstance(currency, str) or not re.fullmatch(r"[A-Z]{3}", currency):
            raise CourseError("CURRENCY_REQUIRED", "Currency normalization requires an explicit three-letter currency in the schema.")
        amount = value
        if amount.startswith(currency):
            amount = amount[len(currency):].strip()
        elif re.match(r"^[A-Za-z]{3}", amount):
            raise ValueError("currency_mismatch")
        # '$' remains ambiguous unless the schema explicitly binds it to the named currency.
        symbol = spec.get("symbol")
        if symbol and amount.startswith(symbol):
            amount = amount[len(symbol):].strip()
        return format(number(amount), "f")
    raise CourseError("INVALID_SCHEMA", "Column types: text, email, date, number, currency.")


def normalize(headers: list[str], rows: list[list[str]], schema: dict | None = None, dedupe: bool = False) -> dict:
    schema = schema or {}
    fields = headers_normalized(headers)
    overrides = schema.get("columns", {})
    if not isinstance(overrides, dict) or set(overrides) - set(fields):
        raise CourseError("INVALID_SCHEMA", "Schema columns must refer to normalized header names.")
    columns = []
    for index, name in enumerate(fields):
        values = [row[index] for row in rows]
        spec = overrides.get(name, {"type": infer(values)})
        if not isinstance(spec, dict) or spec.get("type") not in {"text", "email", "date", "number", "currency"}:
            raise CourseError("INVALID_SCHEMA", "Invalid column type definition.")
        columns.append({"name": name, "source_header": headers[index], "inferred_type": infer(values), **spec,
                        "null_count": sum(v.casefold() in NULLS for v in values)})
    output, errors, seen, duplicates = [], [], {}, []
    for index, row in enumerate(rows, 2):
        normalized = {}
        for value, column in zip(row, columns):
            try:
                normalized[column["name"]] = normalize_cell(value, column)
            except (ValueError, InvalidOperation):
                normalized[column["name"]] = value  # Preserve unparsed evidence; do not manufacture a corrected value.
                errors.append({"row": index, "column": column["name"], "reason": "value_does_not_match_declared_type"})
        key = digest(normalized)
        if key in seen:
            duplicates.append({"row": index, "same_as_row": seen[key]})
            if dedupe:
                continue
        seen.setdefault(key, index)
        output.append(normalized)
    return {"columns": columns, "rows": output, "shape": {"input_rows": len(rows), "output_rows": len(output), "columns": len(fields)},
            "grain": schema.get("grain", "unspecified; confirm what one row represents"), "errors": errors, "duplicates": duplicates,
            "deduplication_applied": dedupe, "assumptions": ["One explicit header row", "English numeric separators", "No date or currency locale guessed"]}


def safe_csv_value(value):
    if value is None:
        return ""
    text = str(value)
    if text.startswith(("=", "+", "-", "@", "\t", "\r")):
        try:
            number(text)
        except (ValueError, InvalidOperation):
            return "'" + text
    return text
